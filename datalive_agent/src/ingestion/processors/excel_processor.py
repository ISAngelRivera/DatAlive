"""
Excel document processor using openpyxl and pandas
"""

import logging
from typing import Dict, Any, Tuple, Union
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    import pandas as pd
except ImportError:
    openpyxl = None
    pd = None

from .base_processor import BaseProcessor

logger = logging.getLogger(__name__)


class ExcelProcessor(BaseProcessor):
    """Excel document processor"""
    
    def __init__(self):
        super().__init__()
        self.supported_extensions = ['.xlsx', '.xls']
        
        if not openpyxl or not pd:
            logger.warning("openpyxl or pandas not installed. Excel processing will use fallback method.")
    
    async def extract_content(
        self,
        source: Union[str, Path, Dict[str, Any]],
        **kwargs
    ) -> Tuple[str, Dict[str, Any]]:
        """
        Extract content and metadata from Excel file
        
        Args:
            source: Path to Excel file
            **kwargs: Additional processing parameters
                - sheets: List[str] = None - Specific sheets to process
                - include_formulas: bool = False - Include formula information
                - max_rows: int = None - Limit rows processed per sheet
                - include_charts: bool = False - Include chart descriptions
        
        Returns:
            Tuple of (content, metadata)
        """
        file_path = Path(source)
        
        if not await self.validate_source(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        sheets = kwargs.get('sheets')
        include_formulas = kwargs.get('include_formulas', False)
        max_rows = kwargs.get('max_rows')
        include_charts = kwargs.get('include_charts', False)
        
        try:
            if pd and openpyxl:
                content, metadata = await self._extract_with_pandas_openpyxl(
                    file_path, sheets, include_formulas, max_rows, include_charts
                )
            else:
                # Fallback method
                content, metadata = await self._extract_fallback(file_path)
            
            # Add basic file metadata
            basic_metadata = self.extract_basic_metadata(file_path)
            metadata.update(basic_metadata)
            
            return self.clean_text(content), metadata
            
        except Exception as e:
            logger.error(f"Error processing Excel file {file_path}: {e}")
            raise
    
    async def _extract_with_pandas_openpyxl(
        self,
        file_path: Path,
        sheets: list = None,
        include_formulas: bool = False,
        max_rows: int = None,
        include_charts: bool = False
    ) -> Tuple[str, Dict[str, Any]]:
        """Extract content using pandas and openpyxl"""
        content_parts = []
        metadata = {
            'extraction_method': 'pandas_openpyxl',
            'sheets': [],
            'has_formulas': False,
            'has_charts': False
        }
        
        # Load workbook with openpyxl for metadata
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            # Extract workbook metadata
            props = workbook.properties
            if props:
                metadata.update({
                    'title': props.title or file_path.stem,
                    'author': props.creator,
                    'subject': props.subject,
                    'description': props.description,
                    'keywords': props.keywords,
                    'created_at': props.created.isoformat() if props.created else None,
                    'modified_at': props.modified.isoformat() if props.modified else None
                })
            
            # Get sheet names
            sheet_names = workbook.sheetnames
            if sheets:
                sheet_names = [name for name in sheet_names if name in sheets]
            
            metadata['total_sheets'] = len(workbook.sheetnames)
            metadata['processed_sheets'] = len(sheet_names)
            
            # Process each sheet
            for sheet_name in sheet_names:
                try:
                    sheet_content, sheet_meta = await self._process_sheet(
                        file_path, sheet_name, workbook[sheet_name],
                        include_formulas, max_rows, include_charts
                    )
                    
                    content_parts.append(f"\n## Sheet: {sheet_name}\n")
                    content_parts.append(sheet_content)
                    
                    metadata['sheets'].append(sheet_meta)
                    
                    if sheet_meta.get('has_formulas'):
                        metadata['has_formulas'] = True
                    if sheet_meta.get('has_charts'):
                        metadata['has_charts'] = True
                        
                except Exception as e:
                    logger.warning(f"Error processing sheet {sheet_name}: {e}")
                    content_parts.append(f"\n## Sheet: {sheet_name}\n[Error processing sheet: {str(e)}]\n")
            
            workbook.close()
            
        except Exception as e:
            logger.error(f"Error loading Excel workbook: {e}")
            raise
        
        return '\n'.join(content_parts), metadata
    
    async def _process_sheet(
        self,
        file_path: Path,
        sheet_name: str,
        worksheet,
        include_formulas: bool = False,
        max_rows: int = None,
        include_charts: bool = False
    ) -> Tuple[str, Dict[str, Any]]:
        """Process a single worksheet"""
        content_parts = []
        sheet_metadata = {
            'name': sheet_name,
            'rows': 0,
            'columns': 0,
            'has_data': False,
            'has_formulas': False,
            'has_charts': False
        }
        
        try:
            # Read sheet data with pandas
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                nrows=max_rows,
                keep_default_na=False
            )
            
            if not df.empty:
                sheet_metadata['has_data'] = True
                sheet_metadata['rows'] = len(df)
                sheet_metadata['columns'] = len(df.columns)
                
                # Convert to text representation
                content_parts.append("Data:")
                content_parts.append(df.to_string(index=False, max_rows=max_rows or 1000))
                
                # Add summary statistics for numeric columns
                numeric_cols = df.select_dtypes(include=['number']).columns
                if len(numeric_cols) > 0:
                    content_parts.append("\nNumeric Summary:")
                    content_parts.append(df[numeric_cols].describe().to_string())
            
            # Check for formulas if requested
            if include_formulas:
                formulas = self._extract_formulas(worksheet, max_rows)
                if formulas:
                    sheet_metadata['has_formulas'] = True
                    content_parts.append("\nFormulas:")
                    content_parts.append('\n'.join(formulas))
            
            # Check for charts if requested
            if include_charts:
                charts = self._extract_chart_info(worksheet)
                if charts:
                    sheet_metadata['has_charts'] = True
                    content_parts.append("\nCharts:")
                    content_parts.append('\n'.join(charts))
            
        except Exception as e:
            logger.warning(f"Error processing sheet data for {sheet_name}: {e}")
            content_parts.append(f"[Error reading sheet data: {str(e)}]")
        
        return '\n'.join(content_parts), sheet_metadata
    
    def _extract_formulas(self, worksheet, max_rows: int = None) -> list:
        """Extract formulas from worksheet"""
        formulas = []
        
        try:
            row_limit = min(worksheet.max_row, max_rows or 1000)
            
            for row in worksheet.iter_rows(max_row=row_limit):
                for cell in row:
                    if cell.data_type == 'f' and cell.value:  # Formula cell
                        formulas.append(f"{cell.coordinate}: {cell.value}")
                        
        except Exception as e:
            logger.warning(f"Error extracting formulas: {e}")
        
        return formulas
    
    def _extract_chart_info(self, worksheet) -> list:
        """Extract chart information from worksheet"""
        charts = []
        
        try:
            if hasattr(worksheet, '_charts'):
                for i, chart in enumerate(worksheet._charts):
                    chart_info = f"Chart {i+1}: {type(chart).__name__}"
                    if hasattr(chart, 'title') and chart.title:
                        chart_info += f" - {chart.title.text if hasattr(chart.title, 'text') else chart.title}"
                    charts.append(chart_info)
                    
        except Exception as e:
            logger.warning(f"Error extracting chart info: {e}")
        
        return charts
    
    async def _extract_fallback(self, file_path: Path) -> Tuple[str, Dict[str, Any]]:
        """Fallback extraction method"""
        logger.warning("Using fallback Excel extraction method")
        
        metadata = {
            'extraction_method': 'fallback',
            'title': file_path.stem
        }
        
        # Simple approach - just indicate Excel content
        content = f"Excel Document: {file_path.name}\n"
        content += "Content extraction requires openpyxl and pandas libraries.\n"
        content += f"File size: {file_path.stat().st_size} bytes"
        
        return content, metadata
    
    def get_sheet_names(self, file_path: Path) -> list:
        """Get list of sheet names without full extraction"""
        try:
            if pd:
                excel_file = pd.ExcelFile(file_path)
                return excel_file.sheet_names
            elif openpyxl:
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()
                return sheet_names
            else:
                return []
                
        except Exception as e:
            logger.error(f"Error getting sheet names from {file_path}: {e}")
            return []
    
    def get_excel_info(self, file_path: Path) -> Dict[str, Any]:
        """Get basic Excel information without full extraction"""
        try:
            info = {
                'file_size': file_path.stat().st_size,
                'sheet_names': self.get_sheet_names(file_path)
            }
            
            if openpyxl:
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                info['total_sheets'] = len(workbook.sheetnames)
                
                # Get metadata
                props = workbook.properties
                if props:
                    info.update({
                        'title': props.title,
                        'author': props.creator,
                        'created_at': props.created.isoformat() if props.created else None,
                        'modified_at': props.modified.isoformat() if props.modified else None
                    })
                
                workbook.close()
            
            return info
            
        except Exception as e:
            logger.error(f"Error getting Excel info for {file_path}: {e}")
            return {'error': str(e)}